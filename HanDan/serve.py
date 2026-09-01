"""
菡萏咖啡．本機服務

用途：
    在本機啟動一個小型 HTTP 伺服器，做四件事：
      1. 提供「菡萏咖啡-台股損益存摺.html」網頁。
      2. 提供 /api/xlsx-prices 端點：開啟網頁時自動呼叫，直接讀「菡萏咖啡.xlsx」
         現價欄位顯示，不打外部 API，速度快也不受連線狀況影響。
      3. 提供 /api/close-prices 端點：只有按下網頁上的「取得盤後收盤價」按鈕才會
         呼叫，即時抓取證交所與櫃買中心的收盤價，並同步寫回 xlsx 的現價欄位，
         作為下次開網頁時 /api/xlsx-prices 顯示的依據。
      4. 提供 /api/xlsx-balance 端點：開啟網頁時自動呼叫，讀「菡萏咖啡.xlsx」
         對應年度分頁 B2 儲存格，作為「帳戶餘額」欄位的初始基準值。

為什麼需要這個伺服器：
    瀏覽器基於安全設計，網頁不能直接啟動本機程式，也不能跨來源直接
    呼叫證交所 API。透過本機伺服器，網頁與 API 屬同一來源（localhost），
    上述兩項限制皆可解除。

執行方式（Windows + uv）：
    uv venv
    .venv\\Scripts\\activate
    uv pip install requests
    python serve.py

    啟動後瀏覽器會自動開啟 http://127.0.0.1:8765
"""

from __future__ import annotations

import ast
import json
import re
import sys
import threading
import traceback
import webbrowser
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlsplit

import fetch_close as fc

HOST = "127.0.0.1"
PORT = 8765
HTML_NAME = "菡萏咖啡-台股損益存摺.html"
XLSX_NAME = "菡萏咖啡.xlsx"
# 使用者另存的手動整理版本；Excel 存檔時會重新計算，公式快取通常是完整的，
# 用來補上 XLSX_NAME 因為 openpyxl 寫入而清空的公式快取。純備援，不存在就略過。
BACKUP_XLSX_NAME = "菡萏咖啡_手操版.xlsx"
# 瀏覽器端手動編輯的鏡像檔。這些資料原本只存在 localStorage，外部程式讀不到，
# 手機版因此看不到使用者在電腦上做的修改；頁面每次寫入 storage 就鏡像一份到
# 這裡，讓 build_web.py／雲端快照能把最新狀態一起帶到手機。
STATE_NAME = "mobile_state.json"
# Firebase 服務帳戶金鑰。放著就會自動把快照推上 Firestore 供手機讀取；
# 檔案不存在時整個雲端同步靜默停用，其餘功能完全不受影響。
# 這把金鑰等同資料庫的完整存取權，已列入 .gitignore，不可提交或外流。
# 前一交易日收盤價的落地檔。只有即時報價引擎給得出這個數字，而開啟頁面走的
# 是 xlsx（現價欄位只存一個值），因此抓價時存一份下來，下次開頁面才有基準
# 可以算「今日損益」。內容只有公開的市場報價，不含持股數量或成本。
PREV_CLOSE_NAME = "prev_closes.json"
SERVICE_ACCOUNT_NAME = "firebase-service-account.json"
FIRESTORE_DOC = "handan/snapshot"
# 網頁每次寫入 storage 都會鏡像一次，逐次上傳沒有意義；短時間內的多次
# 異動合併成一次推送，減少 Firestore 寫入次數。
CLOUD_PUSH_DELAY_SEC = 3.0
# 打包成 PyInstaller exe 後 __file__ 會指向暫存解壓目錄，不是 exe 所在資料夾；
# 這裡改用 sys.executable 找到 exe 本身的位置，讓 xlsx/html 這些資料檔仍然
# 從「exe 旁邊」讀寫，而不是暫存目錄。一般用 python 執行時 sys.frozen 不存在，
# 維持原本以 __file__ 為準。
BASE_DIR = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) \
    else Path(__file__).resolve().parent


def _finalize_prices(prices: dict, iso: str, target_iso: str, before_close: bool,
                     log: list[str], prev_closes: dict | None = None) -> dict:
    """把抓到的價格篩成追蹤清單，組成網頁需要的回傳格式。

    prev_closes 是前一交易日收盤價，只有即時報價引擎提供得出來，用於前端
    計算「今日損益」。走官方每日檔的路徑時會是空的，前端該卡片顯示為無資料。
    """
    wanted = set(fc.ETF_CODES) | set(fc.STOCK_CODE_MAP.values())
    prev = prev_closes or {}
    missing = sorted(wanted - prices.keys())

    # 中間那幾行「某來源缺幾檔、由另一個來源補上」講的是過程，很容易被讀成
    # 「現在還缺著」。最後明講最終結果，才不會讓人以為抓價沒抓完。
    log.append(f"持股 {len(wanted)} 檔全數取得"
               if not missing else
               f"⚠ 持股 {len(wanted)} 檔僅取得 {len(wanted) - len(missing)} 檔，"
               f"未取得：{'、'.join(missing)}")

    return {
        "ok": True,
        "date": iso,
        "targetDate": target_iso,
        "beforeClose": before_close,
        "prices": {c: prices[c] for c in wanted if c in prices},
        "prevCloses": {c: prev[c] for c in wanted if c in prev},
        "missing": missing,
        "unresolved": fc.UNRESOLVED,
        "log": log,
        "fetchedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def collect_prices() -> dict:
    """抓取收盤價並整理為網頁可直接使用的格式。

    已收盤時走「即時報價優先」的快路徑：該引擎只查追蹤清單、通常一秒內
    回應，收盤後給的就是當日收盤價。官方每日檔要收盤後約一小時才產出，
    在那之前抓回來的是前一交易日的資料，實測光上市的全量備援檔就要 45 秒
    （指定日期端點若再連線失敗，重試會把等待拉到數分鐘），而抓回來的內容
    幾乎整批會被即時報價覆蓋掉，並不划算。

    只有在尚未收盤、或即時報價不可用時，才走官方每日檔的完整路徑。
    """
    target, before_close = fc.target_trading_date()
    target_iso = target.strftime("%Y-%m-%d")
    tracked = sorted(set(fc.ETF_CODES) | set(fc.STOCK_CODE_MAP.values()))
    log: list[str] = []

    # --- 快路徑：已收盤時先問即時報價引擎 ---------------------------------
    realtime: dict[str, float] = {}
    prev_closes: dict[str, float] = {}
    if not before_close:
        try:
            rt_prices, rt_prev, rt_date = fc.fetch_realtime_quotes(tracked)
            rt_iso = fc.roc_to_iso(rt_date)
            if rt_prices and rt_iso == target_iso:
                realtime = rt_prices
                prev_closes = rt_prev
                log.append(f"即時報價：{len(realtime)} 檔 {target_iso} 收盤價")
            elif rt_prices:
                log.append(f"即時報價回傳 {rt_iso}，非目標日，改抓官方每日檔")
            else:
                log.append("即時報價查無資料，改抓官方每日檔")
        except Exception as exc:  # noqa: BLE001
            log.append(f"即時報價失敗（{type(exc).__name__}），改抓官方每日檔")

    if realtime:
        # 即時報價涵蓋上市與上櫃，補不到的只剩興櫃（該引擎不提供興櫃報價，
        # tse/otc/emg/rot 四種前綴皆實測無成交價）。因此只補這一組資料集，
        # 跳過最慢的上市／上櫃官方全量檔。
        esb: dict[str, float] = {}
        esb_date = ""
        try:
            esb, _esb_src, esb_date = fc.fetch_tpex_esb()
            log.append(f"興櫃資料集：{len(esb)} 檔" if esb else "興櫃：所有候選端點皆無回應")
        except Exception as exc:  # noqa: BLE001
            log.append(f"興櫃抓取失敗（{type(exc).__name__}）")

        prices = {**esb, **realtime}  # 同一代號以即時報價為準
        if esb and fc.roc_to_iso(esb_date) != target_iso:
            log.append(f"⚠ 興櫃資料為 {fc.roc_to_iso(esb_date)} 收盤價，尚未更新至 {target_iso}")

        # 即時報價對「當日無成交」的商品回 "-"，這類冷門 ETF 會整檔漏掉，
        # 但官方每日檔仍有收盤價（實測 00944 官方檔 20.63、即時報價無成交）。
        # 分兩層補齊，兩層都只在真的有缺漏時才做，不影響正常情況的速度：
        #   1. 上市每日檔只要 0.6 秒，便宜，優先用它補。
        #   2. 上櫃每日檔要 25 秒、而且常常還沒更新到當日，代價不成比例；
        #      改用即時報價已經拿到的前一交易日收盤價頂替。當日無成交，
        #      用昨收當現價與券商軟體一致，今日損益也會正確算成 0。
        missing = [c for c in tracked if c not in prices]
        if missing:
            try:
                twse, twse_raw = fc.fetch_twse_dated(target)
                if twse and fc.roc_to_iso(twse_raw) == target_iso:
                    hit = [c for c in missing if c in twse]
                    for c in hit:
                        prices[c] = twse[c]
                    if hit:
                        log.append(f"{len(hit)} 檔當日無成交，已改由官方每日檔取得")
            except Exception as exc:  # noqa: BLE001
                log.append(f"官方每日檔補抓失敗（{type(exc).__name__}）")

            fallback = [c for c in tracked if c not in prices and c in prev_closes]
            for c in fallback:
                prices[c] = prev_closes[c]
            if fallback:
                log.append(f"{len(fallback)} 檔當日無成交，以前一交易日收盤價計")

        return _finalize_prices(prices, target_iso, target_iso, before_close, log, prev_closes)

    # --- 完整路徑：尚未收盤，或即時報價不可用 -----------------------------
    # 先試指定日期，失敗則退回最新一期。
    twse: dict[str, float] = {}
    raw_date = ""
    try:
        twse, raw_date = fc.fetch_twse_dated(target)
        if twse:
            log.append(f"上市：指定日期 {target:%Y-%m-%d}，{len(twse)} 檔")
    except Exception as exc:  # noqa: BLE001
        log.append(f"上市指定日期失敗（{type(exc).__name__}），改抓最新一期")

    if not twse:
        twse, raw_date = fc.fetch_twse_latest()
        log.append(f"上市：最新一期，{len(twse)} 檔")

    tpex_date = ""
    try:
        tpex, tpex_date = fc.fetch_tpex()
        log.append(f"上櫃：{len(tpex)} 檔")
    except Exception as exc:  # noqa: BLE001
        tpex = {}
        log.append(f"上櫃抓取失敗（{type(exc).__name__}）")

    # 興櫃屬另一組資料集，7907 等 79xx 號段商品需由此取得。
    esb_date = ""
    try:
        esb, esb_src, esb_date = fc.fetch_tpex_esb()
        log.append(f"興櫃資料集：{len(esb)} 檔" if esb else "興櫃：所有候選端點皆無回應")
    except Exception as exc:  # noqa: BLE001
        esb = {}
        log.append(f"興櫃抓取失敗（{type(exc).__name__}）")

    prices = {**twse, **tpex, **esb}
    iso = fc.roc_to_iso(raw_date) or target_iso

    # 官方每日檔要到收盤後一段時間才產出（實測 13:55 仍回前一交易日的資料），
    # 這時「上市」和「上櫃」兩邊都是舊的，彼此一致，光靠來源互比看不出問題，
    # 使用者剛收盤按下按鈕就只會拿到昨天的數字。因此只要已收盤卻拿到舊日期，
    # 就直接用即時報價引擎（收盤當下即反映最後成交價）補上當日價。
    realtime_applied = False
    if not before_close and iso != target_iso:
        try:
            rt_prices, rt_prev, rt_date = fc.fetch_realtime_quotes(tracked)
        except Exception:  # noqa: BLE001
            rt_prices, rt_prev, rt_date = {}, {}, ""
        if rt_prices and fc.roc_to_iso(rt_date) == target_iso:
            prices.update(rt_prices)
            prev_closes = rt_prev
            log.append(f"官方每日檔仍為 {iso}，已用即時報價補上 {len(rt_prices)} 檔 {target_iso} 收盤價")
            iso = target_iso
            realtime_applied = True
        else:
            log.append(f"⚠ 官方每日檔仍為 {iso}，即時報價也查無 {target_iso} 的資料")

    # 上市（TWSE）與上櫃／興櫃（TPEx）是各自獨立的來源，收盤後更新的時間點不一定
    # 一致；只看 TWSE 的日期會誤以為全部資料都是當日的，因此個別比對並提醒使用者。
    # 上面若已整批補過當日價就不必再補一次。
    if not realtime_applied and tpex and fc.roc_to_iso(tpex_date) != iso:
        # 官方批次檔常常收盤後一兩小時才更新，改用上市櫃共用的即時報價引擎
        # （收盤當下就反映最後成交價）補上當天資料，取得不到才顯示過時警告。
        try:
            rt_prices, rt_prev, rt_date = fc.fetch_realtime_quotes(tracked)
        except Exception:  # noqa: BLE001
            rt_prices, rt_prev, rt_date = {}, {}, ""
        if rt_prices and fc.roc_to_iso(rt_date) == iso:
            prices.update(rt_prices)
            prev_closes = rt_prev
            tpex.update(rt_prices)
            log.append(f"上櫃官方批次檔為 {fc.roc_to_iso(tpex_date)}，已改用即時報價補上 {len(rt_prices)} 檔")
        else:
            log.append(f"⚠ 上櫃資料為 {fc.roc_to_iso(tpex_date)} 收盤價，尚未更新至 {iso}，"
                        "上櫃個股／債券型ETF價格可能不是最新收盤價")
    if esb and fc.roc_to_iso(esb_date) != iso:
        log.append(f"⚠ 興櫃資料為 {fc.roc_to_iso(esb_date)} 收盤價，尚未更新至 {iso}")

    return _finalize_prices(prices, iso, target_iso, before_close, log, prev_closes)


def update_xlsx_prices(prices: dict) -> dict:
    """把追蹤中商品的現價寫入「菡萏咖啡.xlsx」對應年度分頁的「現價」欄位。

    只更新商品代號／名稱落在 fc.STOCK_CODE_MAP／fc.ETF_CODES 追蹤範圍內的列
    （即目前持股，判斷方式與 fetch_close.py 抓價時一致）；不在追蹤範圍的列
    （多半是已賣出的舊部位）維持原樣，不會被覆蓋。

    不依賴固定的列號範圍——逐列掃描每個分頁「損益表」表格區（自第 57 列，
    週績效表下方起算）的商品欄，就地比對，這樣使用者在 Excel 裡插入／刪除
    持股列也不會讓寫入位置跑掉。
    """
    xlsx_path = BASE_DIR / XLSX_NAME
    if not xlsx_path.exists():
        return {"ok": False, "error": f"找不到 {XLSX_NAME}"}

    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "本機環境缺少 openpyxl 套件，無法寫入 xlsx"}

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except PermissionError:
        return {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，無法寫入，請先關閉後再試一次"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"讀取失敗：{type(exc).__name__}: {exc}"}

    year = datetime.now().year
    sheet_names = [f"{year}損益表", f"{year}損益表 (可橙)"]
    name_to_code = dict(fc.STOCK_CODE_MAP)

    updated = 0
    missing_sheets = []
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            missing_sheets.append(sheet_name)
            continue
        ws = wb[sheet_name]
        for r in range(57, ws.max_row + 1):
            name = ws.cell(row=r, column=1).value
            if name is None:
                continue
            code = name_to_code.get(str(name).strip(), str(name).strip())
            if code in prices:
                ws.cell(row=r, column=5).value = prices[code]
                updated += 1

    try:
        wb.save(xlsx_path)
    except PermissionError:
        return {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，無法寫入，請先關閉後再試一次"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"儲存失敗：{type(exc).__name__}: {exc}"}

    return {"ok": True, "updatedCount": updated, "missingSheets": missing_sheets}


# ETF 配息分頁裡「1月」～「12月」欄位固定是 G～R（見 row 19 表頭：
# G19=JAN ... R19=DEC），跟每個商品列裡的月配息公式欄一一對應。
ETF_DIV_MONTH_COL = {
    "jan": "G", "feb": "H", "mar": "I", "apr": "J", "may": "K", "jun": "L",
    "jul": "M", "aug": "N", "sep": "O", "oct": "P", "nov": "Q", "dec": "R",
}


def _fmt_num(v: float) -> str:
    """把數字轉成跟使用者手寫註解一致的簡潔格式：整數不帶小數點，
    小數去掉多餘的尾端 0（例如 0.866、1、3.5），不用科學記號。"""
    s = f"{v:.6f}".rstrip("0").rstrip(".")
    return s if s else "0"


def _open_etf_div_cell(acc: str, name: str, month: str):
    """共用的定位邏輯：開啟活頁簿、找到對應年度的存股(ETF)分頁、依商品名稱
    （B 欄）找到列、依月份找到欄（見 ETF_DIV_MONTH_COL）。

    給 write_etf_dividend()／clear_etf_dividend() 共用，回傳
    (wb, xlsx_path, cell, sheet_name) 或 (None, None, None, error_dict)。
    """
    xlsx_path = BASE_DIR / XLSX_NAME
    if not xlsx_path.exists():
        return None, None, None, {"ok": False, "error": f"找不到 {XLSX_NAME}"}

    col = ETF_DIV_MONTH_COL.get(month)
    if col is None:
        return None, None, None, {"ok": False, "error": f"不認得的月份代碼：{month}"}

    try:
        import openpyxl
    except ImportError:
        return None, None, None, {"ok": False, "error": "本機環境缺少 openpyxl 套件，無法寫入 xlsx"}

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except PermissionError:
        return None, None, None, {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，無法寫入，請先關閉後再試一次"}
    except Exception as exc:  # noqa: BLE001
        return None, None, None, {"ok": False, "error": f"讀取失敗：{type(exc).__name__}: {exc}"}

    year = datetime.now().year
    sheet_name = f"{year}存股(ETF)" if acc == "main" else f"{year}存股(ETF) (可橙)"
    if sheet_name not in wb.sheetnames:
        return None, None, None, {"ok": False, "error": f"找不到分頁「{sheet_name}」"}
    ws = wb[sheet_name]

    row = None
    for r in range(2, ws.max_row + 1):
        cell_name = ws.cell(row=r, column=2).value  # B 欄＝商品
        if cell_name is not None and str(cell_name).strip() == name:
            row = r
            break
    if row is None:
        return None, None, None, {"ok": False, "error": f"「{sheet_name}」裡找不到商品「{name}」"}

    return wb, xlsx_path, ws[f"{col}{row}"], sheet_name


def write_etf_dividend(acc: str, name: str, month: str, date_iso: str,
                        ex_div_amount: float, fee: float, amount: float) -> dict:
    """把「存股與配息」新增的一筆 ETF 配息寫進「菡萏咖啡.xlsx」對應年度的
    存股(ETF) 分頁：依商品名稱（B 欄）找到列，依月份找到欄（見
    ETF_DIV_MONTH_COL），把配息金額寫入該儲存格，並在該格插入一則註解
    （格式：除息{除息金額} {M}/{D} 匯入{配息金額}），方便日後在 Excel 裡
    直接看到這筆配息的原始依據，不用回頭查網頁。

    這會覆蓋該儲存格原本的內容（通常是手動填的公式，例如
    `=0.866*3000-10`），改成這次算出來的實際金額——跟「取得盤後收盤價」
    覆蓋現價欄位是同一種做法：以使用者這次的輸入為準。
    """
    wb, xlsx_path, target, sheet_name = _open_etf_div_cell(acc, name, month)
    if wb is None:
        return sheet_name  # 這裡的第四個回傳值其實是 error dict

    try:
        from openpyxl.comments import Comment
        year_num, month_num, day_num = (int(p) for p in date_iso.split("-")[0:3])
    except (ValueError, IndexError):
        return {"ok": False, "error": f"日期格式錯誤：{date_iso}"}

    target.value = amount
    comment_text = f"除息{_fmt_num(ex_div_amount)} {year_num}/{month_num}/{day_num} 匯入{_fmt_num(amount)}"
    target.comment = Comment(comment_text, "菡萏咖啡存摺（自動）")

    try:
        wb.save(xlsx_path)
    except PermissionError:
        return {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，無法寫入，請先關閉後再試一次"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"儲存失敗：{type(exc).__name__}: {exc}"}

    return {"ok": True, "sheet": sheet_name, "cell": target.coordinate, "comment": comment_text}


def clear_etf_dividend(acc: str, name: str, month: str) -> dict:
    """刪除某商品某月已經寫入的配息：清空儲存格的值與註解，回到空白狀態。

    用於使用者輸入錯欄位時的補救；跟 write_etf_dividend() 共用同一套
    定位邏輯（_open_etf_div_cell），差別只在最後是清空而不是寫入。
    """
    wb, xlsx_path, target, sheet_name = _open_etf_div_cell(acc, name, month)
    if wb is None:
        return sheet_name  # 這裡的第四個回傳值其實是 error dict

    target.value = None
    target.comment = None

    try:
        wb.save(xlsx_path)
    except PermissionError:
        return {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，無法寫入，請先關閉後再試一次"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"儲存失敗：{type(exc).__name__}: {exc}"}

    return {"ok": True, "sheet": sheet_name, "cell": target.coordinate}


# 股票配息位置在「{year}損益表」／「{year}損益表 (可橙)」本身（不是另一個分頁），
# 跟持股同一張表：從第 57 列開始（跟 update_xlsx_prices 找持股列一致），每列
# A 欄＝商品名稱，L 欄＝除權息（累計金額）。這張表往下接著是 ETF 區塊（第二個
# 「商品／庫存數量」表頭列，A 欄會是字串「商品」），股票要找的只有表頭到這裡
# 之間的列，掃到「商品」就停止，避免誤觸到 ETF 那邊的列。
STOCK_DIV_COL = "L"


def _safe_eval_arith(expr: str) -> float | None:
    """安全地算出只含數字與 + - * / 的算式（例如 xlsx 裡常見的
    `0.699*1000-10` 或 `27399+3976+12990`），不支援函式呼叫或儲存格參照，
    解析失敗一律回傳 None，不亂猜。"""
    expr = expr.lstrip("=")
    try:
        node = ast.parse(expr, mode="eval").body
    except SyntaxError:
        return None

    def _eval(n):
        if isinstance(n, ast.Constant) and isinstance(n.value, (int, float)):
            return n.value
        if isinstance(n, ast.BinOp) and isinstance(n.op, (ast.Add, ast.Sub, ast.Mult, ast.Div)):
            left, right = _eval(n.left), _eval(n.right)
            if left is None or right is None:
                return None
            if isinstance(n.op, ast.Add):
                return left + right
            if isinstance(n.op, ast.Sub):
                return left - right
            if isinstance(n.op, ast.Mult):
                return left * right
            return left / right
        if isinstance(n, ast.UnaryOp) and isinstance(n.op, ast.USub):
            v = _eval(n.operand)
            return None if v is None else -v
        return None

    return _eval(node)


def _find_stock_div_row(ws) -> dict[str, int]:
    """掃描整張損益表，收集所有「有除權息追蹤」的持股列，回傳
    {商品名稱: 列號} 的對照表。

    這張表裡股票區塊不只一段——開頭一段（第 57 列表頭起）之後接 ETF 區，
    再往下常常還有「借券(出借)」小計，小計之後又接著另一批股票列，
    所以不能靠「遇到某個表頭列就停止」，必須整張表掃完。用 G 欄（損益）
    是否為 `=F{row}-D{row}+L{row}` 這個公式型態來判斷「這列是有除權息
    欄位可寫的持股列」，藉此跳過「借券(出借)」那種只到 F 欄（市值）、
    沒有 G/H/L 欄位的純出借列，以及表頭列、空白列。ETF 列也符合這個
    公式型態，但 ETF 用代號命名（如 0056），跟股票的中文名稱不會撞名，
    不影響依名稱查找。

    同名列（同一檔股票分批買進，各自一列）只留第一筆出現的列號——這是
    既有限制（前端 baseStocks() 挑股票時也是類似取法，不是本次修正
    範圍），並非本函式新引入的問題。
    """
    rows: dict[str, int] = {}
    for r in range(58, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name is None:
            continue
        name = str(name).strip()
        g_val = ws.cell(row=r, column=7).value  # G 欄＝損益
        if not (isinstance(g_val, str) and g_val.startswith("=") and "+L" in g_val):
            continue
        rows.setdefault(name, r)
    return rows


def _read_existing_stock_dividend(acc: str, name: str) -> tuple[float | None, dict | None]:
    """讀 L 欄目前的累計配息金額。優先讀公式快取（data_only=True），
    快取被清空（見 CLAUDE.md 已知副作用）時改讀 BACKUP_XLSX_NAME 的快取，
    兩者都沒有時，嘗試安全解析公式本身（純數字加減乘除，無儲存格參照的
    才解析得出來）。全部都讀不到就回傳 (None, error)，不悄悄當成 0，
    避免使用者不知情下把舊配息蓋掉。

    回傳 (existing_amount, None) 或 (None, error_dict)。
    """
    xlsx_path = BASE_DIR / XLSX_NAME
    if not xlsx_path.exists():
        return None, {"ok": False, "error": f"找不到 {XLSX_NAME}"}

    try:
        import openpyxl
    except ImportError:
        return None, {"ok": False, "error": "本機環境缺少 openpyxl 套件"}

    sheet_name = f"{datetime.now().year}損益表" if acc == "main" else f"{datetime.now().year}損益表 (可橙)"

    # _find_stock_div_row 靠 G 欄「是不是公式字串」判斷列，這只有在
    # data_only=False（讀到公式本身）時才看得到；data_only=True 讀到的是
    # 快取的計算結果（數字），不是公式字串，用來找列一定會找不到。
    # 所以先用 raw 版本把列號定出來，再依序去各版本的活頁簿讀 L 欄的值。
    try:
        wb_raw = openpyxl.load_workbook(xlsx_path, data_only=False)
    except PermissionError:
        return None, {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，暫時無法讀取"}
    except Exception as exc:  # noqa: BLE001
        return None, {"ok": False, "error": f"讀取失敗：{type(exc).__name__}: {exc}"}

    if sheet_name not in wb_raw.sheetnames:
        return None, {"ok": False, "error": f"找不到分頁「{sheet_name}」"}

    rows = _find_stock_div_row(wb_raw[sheet_name])
    row = rows.get(name)
    if row is None:
        return None, {"ok": False, "error": f"「{sheet_name}」股票區找不到商品「{name}」"}

    try:
        wb_cached = openpyxl.load_workbook(xlsx_path, data_only=True)
        value = wb_cached[sheet_name].cell(row=row, column=12).value  # L 欄
    except Exception:  # noqa: BLE001
        value = None
    if isinstance(value, (int, float)):
        return float(value), None

    backup_path = BASE_DIR / BACKUP_XLSX_NAME
    if backup_path.exists():
        try:
            wb_backup_raw = openpyxl.load_workbook(backup_path, data_only=False)
            wb_backup_cached = openpyxl.load_workbook(backup_path, data_only=True)
            if sheet_name in wb_backup_raw.sheetnames:
                backup_rows = _find_stock_div_row(wb_backup_raw[sheet_name])
                backup_row = backup_rows.get(name)
                if backup_row is not None:
                    backup_value = wb_backup_cached[sheet_name].cell(row=backup_row, column=12).value
                    if isinstance(backup_value, (int, float)):
                        return float(backup_value), None
        except Exception:  # noqa: BLE001
            pass

    raw_value = wb_raw[sheet_name].cell(row=row, column=12).value

    if raw_value is None:
        return 0.0, None  # 儲存格原本就是空的，原配息金額視為 0，這是唯一安全的「猜測」
    if isinstance(raw_value, (int, float)):
        return float(raw_value), None
    if isinstance(raw_value, str) and raw_value.startswith("="):
        evaluated = _safe_eval_arith(raw_value)
        if evaluated is not None:
            return evaluated, None

    return None, {
        "ok": False,
        "error": f"「{name}」目前的除權息儲存格（{sheet_name} L{row}）內容無法自動解析：{raw_value!r}，"
                 f"請先在 Excel 裡確認或手動改成純數字後再試一次",
    }


def write_stock_dividend(acc: str, name: str, date_iso: str, ex_div_amount: float,
                          fee: float, this_amount: float) -> dict:
    """股票配息採累加：讀出目前的除權息累計金額（見
    _read_existing_stock_dividend），加上這次新輸入的金額，把新的總額
    （純數字，不是公式）寫回 L 欄；註解則是「附加」而非覆蓋——保留原本
    的內容，換行加上這次的紀錄（除息{除息金額} {M}/{D} 匯入{此次金額}），
    這次金額指的是這一筆的淨收入，不是累加後的總額，跟使用者原本手寫
    的多行註解是同一種記法。
    """
    existing, err = _read_existing_stock_dividend(acc, name)
    if err is not None:
        return err

    xlsx_path = BASE_DIR / XLSX_NAME
    try:
        import openpyxl
        from openpyxl.comments import Comment
    except ImportError:
        return {"ok": False, "error": "本機環境缺少 openpyxl 套件，無法寫入 xlsx"}

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except PermissionError:
        return {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，無法寫入，請先關閉後再試一次"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"讀取失敗：{type(exc).__name__}: {exc}"}

    sheet_name = f"{datetime.now().year}損益表" if acc == "main" else f"{datetime.now().year}損益表 (可橙)"
    if sheet_name not in wb.sheetnames:
        return {"ok": False, "error": f"找不到分頁「{sheet_name}」"}
    ws = wb[sheet_name]
    rows = _find_stock_div_row(ws)
    row = rows.get(name)
    if row is None:
        return {"ok": False, "error": f"「{sheet_name}」股票區找不到商品「{name}」"}

    try:
        year_num, month_num, day_num = (int(p) for p in date_iso.split("-")[0:3])
    except (ValueError, IndexError):
        return {"ok": False, "error": f"日期格式錯誤：{date_iso}"}

    new_total = existing + this_amount
    target = ws.cell(row=row, column=12)  # L 欄
    target.value = new_total

    new_line = f"除息{_fmt_num(ex_div_amount)} {year_num}/{month_num}/{day_num} 匯入{_fmt_num(this_amount)}"
    if target.comment is not None and target.comment.text:
        target.comment.text = target.comment.text.rstrip("\n") + "\n" + new_line
    else:
        target.comment = Comment(new_line, "菡萏咖啡存摺（自動）")

    try:
        wb.save(xlsx_path)
    except PermissionError:
        return {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，無法寫入，請先關閉後再試一次"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"儲存失敗：{type(exc).__name__}: {exc}"}

    return {"ok": True, "sheet": sheet_name, "cell": target.coordinate, "existing": existing, "newTotal": new_total}


def set_stock_dividend_base(acc: str, name: str, new_value: float) -> dict:
    """直接把股票的「原配息金額」（L 欄累計值）改成 new_value，用於手動
    校正跟 Excel 對不上的情況——這不是新增一筆配息事件，只是修正累計基準，
    所以不會動註解，也不會影響「未入帳配息」／帳戶餘額（那些只在「確認」
    新配息或「匯入」時才會變動）。"""
    xlsx_path = BASE_DIR / XLSX_NAME
    if not xlsx_path.exists():
        return {"ok": False, "error": f"找不到 {XLSX_NAME}"}

    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "本機環境缺少 openpyxl 套件，無法寫入 xlsx"}

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except PermissionError:
        return {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，無法寫入，請先關閉後再試一次"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"讀取失敗：{type(exc).__name__}: {exc}"}

    sheet_name = f"{datetime.now().year}損益表" if acc == "main" else f"{datetime.now().year}損益表 (可橙)"
    if sheet_name not in wb.sheetnames:
        return {"ok": False, "error": f"找不到分頁「{sheet_name}」"}
    ws = wb[sheet_name]
    rows = _find_stock_div_row(ws)
    row = rows.get(name)
    if row is None:
        return {"ok": False, "error": f"「{sheet_name}」股票區找不到商品「{name}」"}

    target = ws.cell(row=row, column=12)  # L 欄
    target.value = new_value

    try:
        wb.save(xlsx_path)
    except PermissionError:
        return {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，無法寫入，請先關閉後再試一次"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"儲存失敗：{type(exc).__name__}: {exc}"}

    return {"ok": True, "sheet": sheet_name, "cell": target.coordinate, "newValue": new_value}


def read_xlsx_prices() -> dict:
    """直接從「菡萏咖啡.xlsx」的現價欄位讀取價格，不打證交所／櫃買中心 API。

    給開啟網頁時的自動顯示用：速度快、不受外部連線狀況影響，如實反映使用者
    上次按「取得盤後收盤價」時同步寫回 xlsx 的內容。判斷追蹤範圍與掃描方式
    與 update_xlsx_prices 一致。
    """
    xlsx_path = BASE_DIR / XLSX_NAME
    if not xlsx_path.exists():
        return {"ok": False, "error": f"找不到 {XLSX_NAME}"}

    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "本機環境缺少 openpyxl 套件"}

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except PermissionError:
        return {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，暫時無法讀取"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"讀取失敗：{type(exc).__name__}: {exc}"}

    year = datetime.now().year
    sheet_names = [f"{year}損益表", f"{year}損益表 (可橙)"]
    name_to_code = dict(fc.STOCK_CODE_MAP)
    wanted = set(fc.ETF_CODES) | set(fc.STOCK_CODE_MAP.values())

    prices: dict[str, float] = {}
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r in range(57, ws.max_row + 1):
            name = ws.cell(row=r, column=1).value
            if name is None:
                continue
            code = name_to_code.get(str(name).strip(), str(name).strip())
            if code not in wanted:
                continue
            price = ws.cell(row=r, column=5).value
            if isinstance(price, (int, float)) and price > 0:
                prices[code] = float(price)

    mtime = datetime.fromtimestamp(xlsx_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
    return {
        "ok": True,
        "date": mtime,
        "targetDate": mtime,
        "beforeClose": False,
        "prices": prices,
        "missing": sorted(wanted - prices.keys()),
        "unresolved": fc.UNRESOLVED,
        "log": [f"讀取自 {XLSX_NAME}（檔案更新於 {mtime}）"],
        "fetchedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def _find_pending_dividend(ws, ws_backup=None) -> float | None:
    """在分頁裡找標示「(未入帳)」的儲存格，回傳它左邊那一格的數值。

    這個標籤在持股／借券小計區塊，代表已經算出但還沒實際入帳的配息金額；
    沒有固定的列號（隨持股筆數增減），逐格掃描比寫死列號穩固。左邊那格常是
    公式，快取被清空（見 read_xlsx_balance 註解）時讀到 None，改讀
    ws_backup 同一個座標的值來補。
    """
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "未入帳" in cell.value:
                left = ws.cell(row=cell.row, column=cell.column - 1).value
                if isinstance(left, (int, float)):
                    return float(left)
                if ws_backup is not None:
                    left_backup = ws_backup.cell(row=cell.row, column=cell.column - 1).value
                    if isinstance(left_backup, (int, float)):
                        return float(left_backup)
    return None


def read_xlsx_balance() -> dict:
    """讀取「菡萏咖啡.xlsx」對應年度分頁的帳戶餘額（B2）與未入帳配息基準值。

    A2 是「帳戶餘額」這個標籤本身，B2 才是數值格；使用者目前多半還沒填，
    讀到空值就回傳 null，網頁端會顯示「尚未設定」並讓使用者直接手動輸入覆寫。

    B2 與未入帳配息左邊那格常是公式，寫入「現價」時 openpyxl 會清空全活頁簿
    的公式快取（見 CLAUDE.md 已知副作用），讀到 None 很正常，不代表資料遺失。
    跟 read_xlsx_weekly() 一樣，讀到 None 時改讀 BACKUP_XLSX_NAME（使用者
    另存的手動整理版本，Excel 存檔時已重新計算過，快取通常是完整的）同一格
    的值來補。
    """
    xlsx_path = BASE_DIR / XLSX_NAME
    if not xlsx_path.exists():
        return {"ok": False, "error": f"找不到 {XLSX_NAME}"}

    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "本機環境缺少 openpyxl 套件"}

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except PermissionError:
        return {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，暫時無法讀取"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"讀取失敗：{type(exc).__name__}: {exc}"}

    wb_backup = None
    backup_path = BASE_DIR / BACKUP_XLSX_NAME
    if backup_path.exists():
        try:
            wb_backup = openpyxl.load_workbook(backup_path, data_only=True)
        except Exception:  # noqa: BLE001
            wb_backup = None

    year = datetime.now().year
    sheet_map = {"main": f"{year}損益表", "keqiang": f"{year}損益表 (可橙)"}
    balance: dict[str, float | None] = {}
    pending_dividend: dict[str, float | None] = {}
    for acc, sheet_name in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            balance[acc] = None
            pending_dividend[acc] = None
            continue
        ws = wb[sheet_name]
        ws_backup = wb_backup[sheet_name] if (wb_backup and sheet_name in wb_backup.sheetnames) else None
        value = ws["B2"].value
        if not isinstance(value, (int, float)) and ws_backup is not None:
            value = ws_backup["B2"].value
        balance[acc] = float(value) if isinstance(value, (int, float)) else None
        pending_dividend[acc] = _find_pending_dividend(ws, ws_backup)

    return {"ok": True, "balance": balance, "pendingDividend": pending_dividend}


def _find_pending_dividend_cell(ws):
    """跟 _find_pending_dividend() 找同一個「未入帳」標籤左邊的儲存格，但回傳
    Cell 物件本身（用來寫入），不是數值。"""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "未入帳" in cell.value:
                return ws.cell(row=cell.row, column=cell.column - 1)
    return None


_LAST_NUM_TERM_RE = re.compile(r"[+-]\d+(?:\.\d+)?$")


def _append_delta_to_cell(cell, delta: float, cancel_last: bool = False) -> None:
    """把 delta 疊加到儲存格現有內容上。

    帳戶餘額（B2）、未入帳配息（見 _find_pending_dividend_cell）這兩格，
    使用者原本的記帳習慣是每筆異動都在公式後面接一個 `+數字` 或 `-數字`，
    累積成一長串完整的異動歷史（例如 B2 常是
    `='2025損益表'!E2-L2-K2-M136+M2+890+375+...` 這種形式）。維持這個
    慣例，把新的異動接在後面，而不是整格覆蓋成一個算好的數字，否則會把
    使用者自己一筆一筆記的歷史砍光。原本是空格或已經是純數字（沒有公式
    歷史可保留）時，才直接做數值加總。

    cancel_last=True 時（未入帳配息這格用這個模式）：如果現有公式最後一項
    剛好跟這次的 delta 互相抵銷（例如「確認」剛加了 +190，緊接著「取消／
    匯入」要扣掉同一筆 190），就直接把最後那一項砍掉，而不是再接一項
    `-190` 讓公式留下「+190-190」這種沒意義的一來一回。只比對「最後一項」
    ，不掃整條公式，避免誤刪使用者自己手動記的歷史數字（那些也可能剛好
    跟某次 delta 互相抵銷，但不是我們加的）。
    """
    current = cell.value
    if isinstance(current, str) and current.startswith("="):
        if cancel_last:
            m = _LAST_NUM_TERM_RE.search(current)
            if m:
                try:
                    last_value = float(m.group(0))
                except ValueError:
                    last_value = None
                if last_value is not None and abs(last_value + delta) < 1e-9:
                    cell.value = current[: m.start()]
                    return
        suffix = f"+{delta}" if delta >= 0 else str(delta)
        cell.value = current + suffix
    elif isinstance(current, (int, float)):
        cell.value = current + delta
    else:
        cell.value = delta


def _load_balance_workbooks(xlsx_path):
    """回傳 (wb_cached, wb_backup_cached, error)。wb_cached 用來讀取帳戶餘額／
    未入帳配息「目前」的估計值（寫入前的基準），wb_backup_cached 是備援檔案
    的快取版本，跟 read_xlsx_balance() 用同一套邏輯。"""
    import openpyxl
    try:
        wb_cached = openpyxl.load_workbook(xlsx_path, data_only=True)
    except PermissionError:
        return None, None, {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，無法寫入，請先關閉後再試一次"}
    except Exception as exc:  # noqa: BLE001
        return None, None, {"ok": False, "error": f"讀取失敗：{type(exc).__name__}: {exc}"}

    wb_backup_cached = None
    backup_path = BASE_DIR / BACKUP_XLSX_NAME
    if backup_path.exists():
        try:
            wb_backup_cached = openpyxl.load_workbook(backup_path, data_only=True)
        except Exception:  # noqa: BLE001
            wb_backup_cached = None
    return wb_cached, wb_backup_cached, None


def _read_balance_and_pending(wb_cached, wb_backup_cached, sheet_name: str) -> tuple[float | None, float | None]:
    if sheet_name not in wb_cached.sheetnames:
        return None, None
    ws = wb_cached[sheet_name]
    ws_backup = wb_backup_cached[sheet_name] if (wb_backup_cached and sheet_name in wb_backup_cached.sheetnames) else None
    value = ws["B2"].value
    if not isinstance(value, (int, float)) and ws_backup is not None:
        value = ws_backup["B2"].value
    balance = float(value) if isinstance(value, (int, float)) else None
    pending = _find_pending_dividend(ws, ws_backup)
    return balance, pending


def adjust_pending_dividend(acc: str, delta: float) -> dict:
    """把 delta（可正可負）疊加到「未入帳配息」那格，用附加公式的方式寫入
    （見 _append_delta_to_cell）。回傳的 newPendingDividend 是估計值（讀到的
    快取值＋delta 算出來的），不是重新讀公式精算出來的結果——公式本身常有
    跨分頁參照，存檔後快取又會照例被清空，沒辦法立刻拿到 Excel 真正算出來
    的值，這點跟其他寫入函式一致。
    """
    xlsx_path = BASE_DIR / XLSX_NAME
    if not xlsx_path.exists():
        return {"ok": False, "error": f"找不到 {XLSX_NAME}"}

    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "本機環境缺少 openpyxl 套件，無法寫入 xlsx"}

    year = datetime.now().year
    sheet_name = f"{year}損益表" if acc == "main" else f"{year}損益表 (可橙)"

    wb_cached, wb_backup_cached, err = _load_balance_workbooks(xlsx_path)
    if err is not None:
        return err
    _, pending_before = _read_balance_and_pending(wb_cached, wb_backup_cached, sheet_name)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except PermissionError:
        return {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，無法寫入，請先關閉後再試一次"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"讀取失敗：{type(exc).__name__}: {exc}"}

    if sheet_name not in wb.sheetnames:
        return {"ok": False, "error": f"找不到分頁「{sheet_name}」"}
    ws = wb[sheet_name]
    cell = _find_pending_dividend_cell(ws)
    if cell is None:
        return {"ok": False, "error": f"「{sheet_name}」裡找不到「未入帳」標示欄位"}

    _append_delta_to_cell(cell, delta, cancel_last=True)

    try:
        wb.save(xlsx_path)
    except PermissionError:
        return {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，無法寫入，請先關閉後再試一次"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"儲存失敗：{type(exc).__name__}: {exc}"}

    return {"ok": True, "sheet": sheet_name, "cell": cell.coordinate, "newPendingDividend": (pending_before or 0) + delta}


def import_dividend(acc: str, amount: float) -> dict:
    """「匯入」一筆已登記的配息：未入帳配息扣掉 amount、帳戶餘額加上
    amount，兩格都用附加公式的方式一次寫入、一次存檔（避免存兩次檔、
    公式快取被清空兩次）。"""
    xlsx_path = BASE_DIR / XLSX_NAME
    if not xlsx_path.exists():
        return {"ok": False, "error": f"找不到 {XLSX_NAME}"}

    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "本機環境缺少 openpyxl 套件，無法寫入 xlsx"}

    year = datetime.now().year
    sheet_name = f"{year}損益表" if acc == "main" else f"{year}損益表 (可橙)"

    wb_cached, wb_backup_cached, err = _load_balance_workbooks(xlsx_path)
    if err is not None:
        return err
    balance_before, pending_before = _read_balance_and_pending(wb_cached, wb_backup_cached, sheet_name)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except PermissionError:
        return {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，無法寫入，請先關閉後再試一次"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"讀取失敗：{type(exc).__name__}: {exc}"}

    if sheet_name not in wb.sheetnames:
        return {"ok": False, "error": f"找不到分頁「{sheet_name}」"}
    ws = wb[sheet_name]

    pending_cell = _find_pending_dividend_cell(ws)
    if pending_cell is None:
        return {"ok": False, "error": f"「{sheet_name}」裡找不到「未入帳」標示欄位"}
    balance_cell = ws["B2"]

    _append_delta_to_cell(pending_cell, -amount, cancel_last=True)
    _append_delta_to_cell(balance_cell, amount)

    try:
        wb.save(xlsx_path)
    except PermissionError:
        return {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，無法寫入，請先關閉後再試一次"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"儲存失敗：{type(exc).__name__}: {exc}"}

    return {
        "ok": True,
        "sheet": sheet_name,
        "newPendingDividend": (pending_before or 0) - amount,
        "newBalance": (balance_before or 0) + amount,
    }


def import_stock_dividend(acc: str, name: str, date_iso: str, ex_div_amount: float,
                           fee: float, this_amount: float) -> dict:
    """股票配息的「匯入」：這是真正動到 xlsx 的動作（「確認」只存網頁本地
    狀態＋未入帳配息，不碰股票的除權息累計欄，見前端 openStockDivEntryModal
    的設計說明）。一次做三件事、一次存檔：
      1. 除權息累計（L 欄）加上 this_amount，變成新的總額，並在該儲存格
         附加一行註解（除息{ex_div_amount} {日期} 匯入{this_amount}）——
         這裡才是「除息…匯入…」這個註解措辭真正對應的時刻；
      2. 未入帳配息扣掉 this_amount；
      3. 帳戶餘額加上 this_amount。
    """
    existing, err = _read_existing_stock_dividend(acc, name)
    if err is not None:
        return err

    xlsx_path = BASE_DIR / XLSX_NAME
    try:
        import openpyxl
        from openpyxl.comments import Comment
    except ImportError:
        return {"ok": False, "error": "本機環境缺少 openpyxl 套件，無法寫入 xlsx"}

    year = datetime.now().year
    sheet_name = f"{year}損益表" if acc == "main" else f"{year}損益表 (可橙)"

    wb_cached, wb_backup_cached, load_err = _load_balance_workbooks(xlsx_path)
    if load_err is not None:
        return load_err
    balance_before, pending_before = _read_balance_and_pending(wb_cached, wb_backup_cached, sheet_name)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except PermissionError:
        return {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，無法寫入，請先關閉後再試一次"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"讀取失敗：{type(exc).__name__}: {exc}"}

    if sheet_name not in wb.sheetnames:
        return {"ok": False, "error": f"找不到分頁「{sheet_name}」"}
    ws = wb[sheet_name]

    rows = _find_stock_div_row(ws)
    row = rows.get(name)
    if row is None:
        return {"ok": False, "error": f"「{sheet_name}」股票區找不到商品「{name}」"}

    try:
        year_num, month_num, day_num = (int(p) for p in date_iso.split("-")[0:3])
    except (ValueError, IndexError):
        return {"ok": False, "error": f"日期格式錯誤：{date_iso}"}

    new_total = existing + this_amount
    div_cell = ws.cell(row=row, column=12)  # L 欄
    div_cell.value = new_total
    new_line = f"除息{_fmt_num(ex_div_amount)} {year_num}/{month_num}/{day_num} 匯入{_fmt_num(this_amount)}"
    if div_cell.comment is not None and div_cell.comment.text:
        div_cell.comment.text = div_cell.comment.text.rstrip("\n") + "\n" + new_line
    else:
        div_cell.comment = Comment(new_line, "菡萏咖啡存摺（自動）")

    pending_cell = _find_pending_dividend_cell(ws)
    if pending_cell is None:
        return {"ok": False, "error": f"「{sheet_name}」裡找不到「未入帳」標示欄位"}
    balance_cell = ws["B2"]
    _append_delta_to_cell(pending_cell, -this_amount, cancel_last=True)
    _append_delta_to_cell(balance_cell, this_amount)

    try:
        wb.save(xlsx_path)
    except PermissionError:
        return {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，無法寫入，請先關閉後再試一次"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"儲存失敗：{type(exc).__name__}: {exc}"}

    return {
        "ok": True,
        "sheet": sheet_name,
        "divCell": div_cell.coordinate,
        "newDivTotal": new_total,
        "newPendingDividend": (pending_before or 0) - this_amount,
        "newBalance": (balance_before or 0) + this_amount,
    }


def read_xlsx_weekly() -> dict:
    """讀取「菡萏咖啡.xlsx」當年損益表 A4:I55 的每週記帳資料，取代網頁原本用
    PORTFOLIO_DATA 靜態資料＋JS 計算週績效的方式，改以 xlsx 為當年度的準。

    週績效（I 欄「與前一周比」＝本週累計績效−上週累計績效）不直接讀 I 欄的
    快取值——寫入「現價」「帳戶餘額」時 openpyxl 會清空全活頁簿的公式快取
    （已知副作用，見 CLAUDE.md），I 欄常常因此讀到空值。改成只讀 B（總成本）
    C（即時庫存）這兩欄，用同一套規則（本週累計績效−上週累計績效）重新算
    一次，數字跟 I 欄公式應得結果一致，但不依賴 Excel 的計算快取。

    B／C 本身也可能是公式（例如可橙戶頭整欄 B 是 =$B$1，我的戶頭的 B1 是
    ='2025損益表'!B1-100000），一樣會被快取清空影響。單一儲存格讀到 None
    時，改讀 BACKUP_XLSX_NAME（使用者另存的手動整理版本，Excel 存檔時已
    重新計算過，快取通常是完整的）同一格的值來補。
    """
    xlsx_path = BASE_DIR / XLSX_NAME
    if not xlsx_path.exists():
        return {"ok": False, "error": f"找不到 {XLSX_NAME}"}

    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "本機環境缺少 openpyxl 套件"}

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except PermissionError:
        return {"ok": False, "error": f"{XLSX_NAME} 目前在其他程式（可能是 Excel）中開啟，暫時無法讀取"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"讀取失敗：{type(exc).__name__}: {exc}"}

    wb_backup = None
    backup_path = BASE_DIR / BACKUP_XLSX_NAME
    if backup_path.exists():
        try:
            wb_backup = openpyxl.load_workbook(backup_path, data_only=True)
        except Exception:  # noqa: BLE001
            wb_backup = None

    def resolve_cell(ws, ws_backup, coord: str):
        value = ws[coord].value
        if value is None and ws_backup is not None:
            value = ws_backup[coord].value
        return value

    year = datetime.now().year
    sheet_map = {"main": f"{year}損益表", "keqiang": f"{year}損益表 (可橙)"}
    prev_sheet_map = {"main": f"{year - 1}損益表", "keqiang": f"{year - 1}損益表 (可橙)"}
    weekly: dict[str, list[dict]] = {"main": [], "keqiang": []}

    for acc, sheet_name in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws_backup = wb_backup[sheet_name] if (wb_backup and sheet_name in wb_backup.sheetnames) else None
        resolved = lambda coord: resolve_cell(ws, ws_backup, coord)  # noqa: E731

        b1 = resolved("B1")  # 股本；部分分頁每列的 B 欄是公式（=$B$1 或跨年度公式），快取清空時退回讀這裡。
        rows = []
        for r in range(4, 56):
            date_cell = ws.cell(row=r, column=1).value
            inventory = resolved(f"C{r}")
            if date_cell is None or inventory is None:
                continue  # 還沒記帳的週次
            cost = resolved(f"B{r}")
            if not isinstance(cost, (int, float)) or cost == 0:
                cost = b1 if isinstance(b1, (int, float)) and b1 else None
            if not cost or not isinstance(inventory, (int, float)):
                continue
            try:
                date_iso = date_cell.strftime("%Y-%m-%d")
            except AttributeError:
                continue
            total_pnl = inventory - cost
            rows.append({
                "date": date_iso, "cost": cost, "inventory": inventory,
                "total_pnl": total_pnl, "cum_pct": total_pnl / cost,
            })
        rows.sort(key=lambda x: x["date"])

        # 每年第一週的「與前一周比」不是跟前一週比（沒有前一週），原始公式是
        # =G4/B4，其中 G4 = F4 - 前一年損益表!F55（今年第一週損益 − 去年
        # 最後一週損益），也就是跟去年底的績效銜接。用同一套 resolved() 邏輯
        # 讀去年分頁第 55 列算出去年年底損益，重現這個特例；讀不到就維持
        # None（沒有比較基準），不強行湊數字。
        prev_year_end_pnl = None
        prev_sheet_name = prev_sheet_map[acc]
        if prev_sheet_name in wb.sheetnames:
            ws_prev = wb[prev_sheet_name]
            ws_prev_backup = wb_backup[prev_sheet_name] if (wb_backup and prev_sheet_name in wb_backup.sheetnames) else None
            prev_cost = resolve_cell(ws_prev, ws_prev_backup, "B55")
            prev_inv = resolve_cell(ws_prev, ws_prev_backup, "C55")
            if isinstance(prev_cost, (int, float)) and isinstance(prev_inv, (int, float)):
                prev_year_end_pnl = prev_inv - prev_cost

        for i, row in enumerate(rows):
            if i == 0:
                row["week_pnl"] = (row["total_pnl"] - prev_year_end_pnl) if prev_year_end_pnl is not None else None
                row["wow_pct"] = (row["week_pnl"] / row["cost"]) if row["week_pnl"] is not None else None
            else:
                row["week_pnl"] = row["total_pnl"] - rows[i - 1]["total_pnl"]
                row["wow_pct"] = row["cum_pct"] - rows[i - 1]["cum_pct"]
        weekly[acc] = rows

    return {"ok": True, "year": year, "weekly": weekly}


def save_browser_state(payload: dict) -> dict:
    """把瀏覽器 localStorage 的內容鏡像存成 mobile_state.json。

    頁面每次寫入 storage 後會呼叫這支 API。手機版是另一個瀏覽器、另一份
    localStorage，看不到電腦上的編輯，這份鏡像就是雲端快照唯一拿得到
    使用者手動修改的來源。

    寫檔採「先寫暫存再取代」，避免建置腳本剛好讀到寫到一半的檔案。
    """
    data = payload.get("data")
    if not isinstance(data, dict):
        return {"ok": False, "error": "payload.data 必須是物件"}

    # 空狀態不覆蓋既有鏡像。用另一個瀏覽器、無痕視窗或剛清過快取的環境開啟
    # 頁面時，localStorage 是空的，若照寫就會把先前正確的編輯內容清掉。
    # 真的要清空請直接刪除 mobile_state.json。
    path_existing = BASE_DIR / STATE_NAME
    if not data and path_existing.exists():
        return {"ok": True, "skipped": True, "reason": "空狀態，保留既有鏡像"}

    out = {
        "savedAt": payload.get("savedAt") or datetime.now().isoformat(timespec="seconds"),
        "count": len(data),
        "data": data,
    }
    path = BASE_DIR / STATE_NAME
    tmp = path.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(out, ensure_ascii=False), encoding="utf-8")
    tmp.replace(path)
    return {"ok": True, "count": len(data), "savedAt": out["savedAt"]}


def read_browser_state() -> dict:
    """讀回 mobile_state.json；不存在或損毀時回傳未設定狀態，不拋錯。"""
    path = BASE_DIR / STATE_NAME
    if not path.exists():
        return {"ok": False, "error": "尚未有任何鏡像資料"}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"{type(exc).__name__}: {exc}"}
    data["ok"] = True
    return data


def save_prev_closes(price_date: str, prev_closes: dict) -> None:
    """把前一交易日收盤價落地，供之後開啟頁面時計算今日損益。

    一併記下這批價格的交易日（price_date），前端才能判斷基準是否還適用：
    隔天若沒重新抓價，這份資料就過期了，卡片上會顯示舊的基準日提醒使用者。
    空的就不寫，避免把可用的舊基準蓋掉。
    """
    if not prev_closes:
        return
    path = BASE_DIR / PREV_CLOSE_NAME
    tmp = path.with_suffix(".json.tmp")
    payload = {"ok": True, "priceDate": price_date, "prevCloses": prev_closes}
    tmp.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    tmp.replace(path)


def read_prev_closes() -> dict:
    """讀回落地的前一交易日收盤價；不存在或損毀時回傳未設定狀態，不拋錯。"""
    path = BASE_DIR / PREV_CLOSE_NAME
    if not path.exists():
        return {"ok": False, "error": "尚未有前一交易日收盤價，請按一次「取得盤後收盤價」"}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"{type(exc).__name__}: {exc}"}


def read_prices_with_basis() -> dict:
    """讀 xlsx 現價，並附上今日損益需要的前一交易日基準。

    /api/xlsx-prices 與手機／雲端版的快照都走這一支，兩邊結構才會一致。
    少了基準的話手機上的今日損益卡片會一直顯示「需按一次取得盤後收盤價」，
    而手機根本按不了那個按鈕。
    """
    data = read_xlsx_prices()
    stored = read_prev_closes()
    if stored.get("ok"):
        data["prevCloses"] = stored.get("prevCloses") or {}
        data["prevCloseDate"] = stored.get("priceDate") or ""
    return data


def read_portfolio_data() -> dict:
    """從主檔 HTML 取出 PORTFOLIO_DATA 這份歷史損益基準資料。

    Artifact 版是私人頁面，可以直接把這份資料內嵌在檔案裡；但部署到 Vercel
    的雲端版網址是公開的，內嵌等於讓任何人不必登入、看網頁原始碼就能取得
    2021 年至今的完整帳本。因此雲端版的 HTML 不含這段，改由這裡讀出來連同
    快照一起推上 Firestore，登入後才下載。
    """
    html_path = BASE_DIR / HTML_NAME
    if not html_path.exists():
        return {"ok": False, "error": f"找不到 {HTML_NAME}"}

    text = html_path.read_text(encoding="utf-8")
    match = re.search(r"const PORTFOLIO_DATA = (\{.*?\});", text, re.S)
    if not match:
        return {"ok": False, "error": "HTML 中找不到 PORTFOLIO_DATA 區塊"}

    try:
        return {"ok": True, "data": json.loads(match.group(1))}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"{type(exc).__name__}: {exc}"}


def read_stock_code_map() -> dict:
    """從主檔 HTML 取出個股名稱對照表。

    這份表列出所有持股名稱與代號，等於把投資組合攤開來，因此雲端版同樣
    不內嵌，改由登入後注入。刻意解析 HTML 裡那一份、而不是直接用
    fetch_close.STOCK_CODE_MAP，因為兩邊允許有差異（例如同一代號的舊稱與
    新稱並存），這裡要的是畫面實際使用的那一份。
    """
    html_path = BASE_DIR / HTML_NAME
    if not html_path.exists():
        return {"ok": False, "error": f"找不到 {HTML_NAME}"}

    text = html_path.read_text(encoding="utf-8")
    match = re.search(r"const STOCK_CODE_MAP = (\{.*?\n\});", text, re.S)
    if not match:
        return {"ok": False, "error": "HTML 中找不到 STOCK_CODE_MAP 區塊"}

    try:
        # JS 的物件字面量在這裡剛好也是合法的 Python 字面量（單引號、
        # 允許尾隨逗號），用 literal_eval 解析比自己寫剖析器安全。
        return {"ok": True, "data": ast.literal_eval(match.group(1))}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"{type(exc).__name__}: {exc}"}


def build_snapshot() -> dict:
    """彙整手機端需要的全部唯讀資料。

    任一項讀取失敗都不中斷：該項標記 ok=False，手機端會自行退回可用的
    替代來源，行為與讀不到資料時一致。
    """
    snapshot: dict[str, dict] = {}
    for key, fn in (("prices", read_prices_with_basis),
                    ("balance", read_xlsx_balance),
                    ("weekly", read_xlsx_weekly)):
        try:
            snapshot[key] = fn()
        except Exception as exc:  # noqa: BLE001
            snapshot[key] = {"ok": False, "error": f"{type(exc).__name__}: {exc}"}
    snapshot["state"] = read_browser_state()
    snapshot["portfolio"] = read_portfolio_data()
    snapshot["codes"] = read_stock_code_map()
    return snapshot


_cloud_session = None
_cloud_project = ""
_cloud_ready: bool | None = None  # None 表示尚未嘗試初始化
_cloud_timer: threading.Timer | None = None
_cloud_lock = threading.Lock()


def _init_cloud() -> bool:
    """準備 Firestore 連線。金鑰不存在或套件未安裝時回傳 False 並停用同步。"""
    global _cloud_session, _cloud_project, _cloud_ready
    if _cloud_ready is not None:
        return _cloud_ready

    key_path = BASE_DIR / SERVICE_ACCOUNT_NAME
    if not key_path.exists():
        print(f"[雲端同步] 未找到 {SERVICE_ACCOUNT_NAME}，略過雲端同步（其餘功能不受影響）")
        _cloud_ready = False
        return False

    try:
        from google.auth.transport.requests import AuthorizedSession
        from google.oauth2 import service_account

        info = json.loads(key_path.read_text(encoding="utf-8"))
        _cloud_project = info["project_id"]
        cred = service_account.Credentials.from_service_account_file(
            str(key_path), scopes=["https://www.googleapis.com/auth/datastore"])
        _cloud_session = AuthorizedSession(cred)
        print(f"[雲端同步] 已啟用，專案 {_cloud_project}")
        _cloud_ready = True
    except Exception as exc:  # noqa: BLE001
        print(f"[雲端同步] 初始化失敗（{type(exc).__name__}: {exc}），略過雲端同步")
        _cloud_ready = False
    return _cloud_ready


def push_snapshot_to_cloud() -> dict:
    """把最新快照寫進 Firestore，供手機端讀取。

    整包快照序列化成單一字串欄位，而不是展開成 Firestore 的巢狀型別。
    這樣不必為每個欄位做型別包裝，前端拿到後 JSON.parse 即可，兩邊的
    資料結構也自然與本機 API 回傳的完全一致。
    """
    if not _init_cloud():
        return {"ok": False, "error": "雲端同步未啟用"}

    payload = json.dumps(build_snapshot(), ensure_ascii=False)
    url = (f"https://firestore.googleapis.com/v1/projects/{_cloud_project}"
           f"/databases/(default)/documents/{FIRESTORE_DOC}")
    body = {"fields": {
        "payload": {"stringValue": payload},
        "updatedAt": {"stringValue": datetime.now().isoformat(timespec="seconds")},
    }}
    try:
        r = _cloud_session.patch(url, json=body, timeout=30)
    except Exception as exc:  # noqa: BLE001
        print(f"[雲端同步] 推送失敗（{type(exc).__name__}: {exc}）")
        return {"ok": False, "error": str(exc)}

    if r.status_code != 200:
        print(f"[雲端同步] 推送失敗 HTTP {r.status_code}：{r.text[:200]}")
        return {"ok": False, "error": f"HTTP {r.status_code}"}

    print(f"[雲端同步] 已更新（{len(payload):,} bytes）")
    return {"ok": True, "bytes": len(payload)}


def schedule_cloud_push(delay: float = CLOUD_PUSH_DELAY_SEC) -> None:
    """延遲推送，短時間內的多次呼叫只會實際送出最後一次。"""
    global _cloud_timer
    if _cloud_ready is False:
        return
    with _cloud_lock:
        if _cloud_timer is not None:
            _cloud_timer.cancel()
        _cloud_timer = threading.Timer(delay, push_snapshot_to_cloud)
        _cloud_timer.daemon = True
        _cloud_timer.start()


class Handler(BaseHTTPRequestHandler):
    """處理網頁與 API 請求。"""

    def _send(self, status: int, body: bytes, ctype: str) -> None:
        self.send_response(status)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def _send_json(self, status: int, payload: dict) -> None:
        self._send(status, json.dumps(payload, ensure_ascii=False).encode("utf-8"),
                   "application/json; charset=utf-8")
        # 資料有異動就排程一次雲端同步。所有寫入類端點都是 POST，唯一的例外
        # 是帶 updateXlsx=1 的取價（GET），因此一併納入。集中在這裡判斷，
        # 日後新增寫入端點不必記得補上推送。
        if status == 200 and (self.command == "POST" or "updateXlsx=1" in self.path):
            schedule_cloud_push()

    def do_GET(self) -> None:  # noqa: N802
        parsed = urlsplit(self.path)
        path = parsed.path
        query = parse_qs(parsed.query)

        if path == "/api/close-prices":
            try:
                result = collect_prices()
                # 前一交易日收盤價只有這條路徑取得到，落地一份供之後開啟頁面
                # 計算今日損益（那時走的是 xlsx，只有現價可讀）。
                save_prev_closes(result.get("date", ""), result.get("prevCloses") or {})
                # 只有網頁上手動按「取得盤後收盤價」才會帶這個參數；開啟網頁時
                # 自動抓價的那一次不寫入 xlsx，避免每次開網頁都動到 Excel 檔案。
                if query.get("updateXlsx", ["0"])[0] == "1":
                    result["xlsx"] = update_xlsx_prices(result["prices"])
                self._send_json(200, result)
            except Exception as exc:  # noqa: BLE001
                traceback.print_exc()
                self._send_json(500, {
                    "ok": False,
                    "error": f"{type(exc).__name__}: {exc}",
                })
            return

        if path == "/api/xlsx-prices":
            try:
                self._send_json(200, read_prices_with_basis())
            except Exception as exc:  # noqa: BLE001
                traceback.print_exc()
                self._send_json(500, {
                    "ok": False,
                    "error": f"{type(exc).__name__}: {exc}",
                })
            return

        if path == "/api/xlsx-balance":
            try:
                self._send_json(200, read_xlsx_balance())
            except Exception as exc:  # noqa: BLE001
                traceback.print_exc()
                self._send_json(500, {
                    "ok": False,
                    "error": f"{type(exc).__name__}: {exc}",
                })
            return

        if path == "/api/xlsx-weekly":
            try:
                self._send_json(200, read_xlsx_weekly())
            except Exception as exc:  # noqa: BLE001
                traceback.print_exc()
                self._send_json(500, {
                    "ok": False,
                    "error": f"{type(exc).__name__}: {exc}",
                })
            return

        if path == "/api/read-stock-dividend":
            try:
                acc = query.get("acc", [""])[0]
                name = query.get("name", [""])[0]
                existing, err = _read_existing_stock_dividend(acc, name)
                self._send_json(200, err if err is not None else {"ok": True, "existingAmount": existing})
            except Exception as exc:  # noqa: BLE001
                traceback.print_exc()
                self._send_json(500, {
                    "ok": False,
                    "error": f"{type(exc).__name__}: {exc}",
                })
            return

        if path in ("/", f"/{HTML_NAME}"):
            html = BASE_DIR / HTML_NAME
            if not html.exists():
                self._send(404, f"找不到 {HTML_NAME}".encode("utf-8"),
                           "text/plain; charset=utf-8")
                return
            self._send(200, html.read_bytes(), "text/html; charset=utf-8")
            return

        self._send(404, b"Not Found", "text/plain; charset=utf-8")

    def do_POST(self) -> None:  # noqa: N802
        parsed = urlsplit(self.path)
        path = parsed.path

        if path == "/api/save-state":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                body = json.loads(self.rfile.read(length) or b"{}")
                self._send_json(200, save_browser_state(body))
            except Exception as exc:  # noqa: BLE001
                traceback.print_exc()
                self._send_json(500, {
                    "ok": False,
                    "error": f"{type(exc).__name__}: {exc}",
                })
            return

        if path == "/api/write-dividend":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                body = json.loads(self.rfile.read(length) or b"{}")
                result = write_etf_dividend(
                    acc=body.get("acc"),
                    name=body.get("name"),
                    month=body.get("month"),
                    date_iso=body.get("date"),
                    ex_div_amount=float(body.get("exDivAmount")),
                    fee=float(body.get("fee") or 0),
                    amount=float(body.get("amount")),
                )
                self._send_json(200, result)
            except Exception as exc:  # noqa: BLE001
                traceback.print_exc()
                self._send_json(500, {
                    "ok": False,
                    "error": f"{type(exc).__name__}: {exc}",
                })
            return

        if path == "/api/clear-dividend":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                body = json.loads(self.rfile.read(length) or b"{}")
                result = clear_etf_dividend(
                    acc=body.get("acc"),
                    name=body.get("name"),
                    month=body.get("month"),
                )
                self._send_json(200, result)
            except Exception as exc:  # noqa: BLE001
                traceback.print_exc()
                self._send_json(500, {
                    "ok": False,
                    "error": f"{type(exc).__name__}: {exc}",
                })
            return

        if path == "/api/write-stock-dividend":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                body = json.loads(self.rfile.read(length) or b"{}")
                result = write_stock_dividend(
                    acc=body.get("acc"),
                    name=body.get("name"),
                    date_iso=body.get("date"),
                    ex_div_amount=float(body.get("exDivAmount")),
                    fee=float(body.get("fee") or 0),
                    this_amount=float(body.get("thisAmount")),
                )
                self._send_json(200, result)
            except Exception as exc:  # noqa: BLE001
                traceback.print_exc()
                self._send_json(500, {
                    "ok": False,
                    "error": f"{type(exc).__name__}: {exc}",
                })
            return

        if path == "/api/set-stock-dividend-base":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                body = json.loads(self.rfile.read(length) or b"{}")
                result = set_stock_dividend_base(
                    acc=body.get("acc"),
                    name=body.get("name"),
                    new_value=float(body.get("newValue")),
                )
                self._send_json(200, result)
            except Exception as exc:  # noqa: BLE001
                traceback.print_exc()
                self._send_json(500, {
                    "ok": False,
                    "error": f"{type(exc).__name__}: {exc}",
                })
            return

        if path == "/api/import-stock-dividend":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                body = json.loads(self.rfile.read(length) or b"{}")
                result = import_stock_dividend(
                    acc=body.get("acc"),
                    name=body.get("name"),
                    date_iso=body.get("date"),
                    ex_div_amount=float(body.get("exDivAmount")),
                    fee=float(body.get("fee") or 0),
                    this_amount=float(body.get("thisAmount")),
                )
                self._send_json(200, result)
            except Exception as exc:  # noqa: BLE001
                traceback.print_exc()
                self._send_json(500, {
                    "ok": False,
                    "error": f"{type(exc).__name__}: {exc}",
                })
            return

        if path == "/api/adjust-pending-dividend":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                body = json.loads(self.rfile.read(length) or b"{}")
                result = adjust_pending_dividend(
                    acc=body.get("acc"),
                    delta=float(body.get("delta")),
                )
                self._send_json(200, result)
            except Exception as exc:  # noqa: BLE001
                traceback.print_exc()
                self._send_json(500, {
                    "ok": False,
                    "error": f"{type(exc).__name__}: {exc}",
                })
            return

        if path == "/api/import-dividend":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                body = json.loads(self.rfile.read(length) or b"{}")
                result = import_dividend(
                    acc=body.get("acc"),
                    amount=float(body.get("amount")),
                )
                self._send_json(200, result)
            except Exception as exc:  # noqa: BLE001
                traceback.print_exc()
                self._send_json(500, {
                    "ok": False,
                    "error": f"{type(exc).__name__}: {exc}",
                })
            return

        self._send(404, b"Not Found", "text/plain; charset=utf-8")

    def log_message(self, fmt: str, *args) -> None:
        """精簡輸出，只保留 API 呼叫紀錄。"""
        if "/api/" in (args[0] if args else ""):
            sys.stderr.write(f"[{datetime.now():%H:%M:%S}] {fmt % args}\n")


def main() -> int:
    html = BASE_DIR / HTML_NAME
    if not html.exists():
        print(f"[錯誤] 找不到 {HTML_NAME}，請確認與 serve.py 放在同一資料夾。")
        return 1

    url = f"http://{HOST}:{PORT}"
    print("=" * 60)
    print("菡萏咖啡．本機服務已啟動")
    print(f"  網址：{url}")
    print("  按下網頁上的「取得盤後收盤價」即會即時抓取。")
    print("  結束服務請按 Ctrl+C。")
    print("=" * 60)

    threading.Timer(1.0, lambda: webbrowser.open(url)).start()

    # 使用多執行緒：抓價需數秒，單執行緒會阻塞網頁本身的請求。
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n[結束] 服務已停止。")
    finally:
        server.server_close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
